import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Alignment, PatternFill

# 支持的图片格式（新增.tif，适配影像文件）
SUPPORT_FORMATS = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif')


class ImageCheckGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("图片质检工具 - 不合格自动记录Excel（断点续检版）")  # 修改标题
        self.root.geometry("900x700")

        # 初始化变量
        self.folder_path = ""
        self.image_list = []
        self.current_index = 0
        self.checked_set = set()
        self.excel_path = ""
        self.photo = None

        # 构建界面
        self._build_ui()

    def _build_ui(self):
        """构建图形界面"""
        # 顶部操作栏
        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.pack(fill=tk.X)

        self.choose_btn = ttk.Button(top_frame, text="选择质检文件夹", command=self.choose_folder)
        self.choose_btn.grid(row=0, column=0, padx=5)
        self.folder_label = ttk.Label(top_frame, text="未选择文件夹")
        self.folder_label.grid(row=0, column=1, padx=5)
        self.status_label = ttk.Label(top_frame, text="状态：等待选择文件夹")
        self.status_label.grid(row=0, column=2, padx=5)

        # 中间图片预览区
        mid_frame = ttk.Frame(self.root, padding="10")
        mid_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(mid_frame, bg="#f5f5f5")
        self.scroll_y = ttk.Scrollbar(mid_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scroll_x = ttk.Scrollbar(mid_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.scroll_y.set, xscrollcommand=self.scroll_x.set)
        self.scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.image_label = tk.Label(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.image_label, anchor=tk.NW)
        self.canvas.bind("<Configure>", self._resize_image)

        # 底部文件名+操作按钮栏
        bottom_frame = ttk.Frame(self.root, padding="10")
        bottom_frame.pack(fill=tk.X)

        self.file_name_label = ttk.Label(bottom_frame, text="当前图片：无", font=("宋体", 12, "bold"))
        self.file_name_label.pack(pady=5)

        # 操作按钮
        self.btn_frame = ttk.Frame(bottom_frame)
        self.btn_frame.pack(pady=5)
        self.prev_btn = ttk.Button(self.btn_frame, text="上一张", command=self.prev_image, state=tk.DISABLED, width=10)
        self.qual_btn = ttk.Button(self.btn_frame, text="合格", command=self.mark_qualified, state=tk.DISABLED,
                                   width=10, style="Accent.TButton")
        self.unqual_btn = ttk.Button(self.btn_frame, text="不合格", command=self.mark_unqualified, state=tk.DISABLED,
                                     width=10, style="Error.TButton")
        self.prev_btn.grid(row=0, column=0, padx=10)
        self.qual_btn.grid(row=0, column=1, padx=10)
        self.unqual_btn.grid(row=0, column=2, padx=10)

        # 按钮样式
        self.root.style = ttk.Style()
        self.root.style.configure("Error.TButton", foreground="red", font=("宋体", 10, "bold"))
        self.root.style.configure("Accent.TButton", foreground="green", font=("宋体", 10, "bold"))

    def _resize_image(self, event=None):
        """自适应缩放图片"""
        if not self.folder_path or self.current_index >= len(self.image_list) or self.current_index < 0:
            self.image_label.config(image="", text="暂无图片")
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            return
        try:
            image_path = self.image_list[self.current_index]
            with Image.open(image_path) as img:
                img = img.convert("RGB")
                canvas_w = self.canvas.winfo_width() - 30
                canvas_h = self.canvas.winfo_height() - 30
                if canvas_w <= 0 or canvas_h <= 0:
                    return
                # 兼容不同PIL版本的缩放参数
                resize_method = Image.ANTIALIAS if hasattr(Image, 'ANTIALIAS') else Image.LANCZOS
                img.thumbnail((canvas_w, canvas_h), resize_method)
                self.photo = ImageTk.PhotoImage(img)
                self.image_label.config(image=self.photo, text="")
                self.canvas.itemconfig(self.canvas_window, width=self.photo.width(), height=self.photo.height())
                self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
        except Exception as e:
            self.image_label.config(image="", text=f"图片加载失败：{str(e)[:50]}")
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            print(f"加载失败详情：{image_path} - {e}")

    def _get_last_unqual_name(self):
        """读取Excel中最后一条不合格记录的文件名"""
        if not os.path.exists(self.excel_path):
            return ""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            if ws.max_row < 2:  # 仅表头
                wb.close()
                return ""
            # 获取最后一行不合格记录
            last_name = ws.cell(row=ws.max_row, column=1).value
            wb.close()
            return last_name.strip() if last_name else ""
        except Exception as e:
            messagebox.showwarning("警告", f"读取历史质检记录失败，将从头开始：{str(e)[:30]}")
            print(f"读取Excel记录详情：{e}")
            return ""

    def choose_folder(self):
        """选择文件夹，初始化图片列表和Excel（断点续检：基于不合格记录）"""
        self.folder_path = filedialog.askdirectory(title="选择要质检的图片文件夹")
        if not self.folder_path:
            messagebox.showinfo("提示", "未选择任何文件夹")
            return

        # 筛选图片文件（小写后缀匹配）
        self.image_list = []
        for f in os.listdir(self.folder_path):
            file_path = os.path.join(self.folder_path, f)
            if os.path.isfile(file_path) and f.lower().endswith(SUPPORT_FORMATS):
                self.image_list.append(file_path)
        self.image_list.sort()  # 固定质检顺序

        if not self.image_list:
            messagebox.showwarning("警告", f"所选文件夹内无支持的图片文件！\n支持格式：{','.join(SUPPORT_FORMATS)}")
            self.folder_label.config(text="未选择文件夹")
            self.status_label.config(text="状态：无图片可质检")
            self.image_list = []
            self.folder_path = ""
            return

        # 初始化Excel（记录不合格图片）
        self.excel_path = os.path.join(self.folder_path, "不合格图片记录.xlsx")
        self._init_excel()

        # 断点续检核心逻辑：从最后一条不合格记录的下一个开始
        self.current_index = 0
        last_unqual_name = self._get_last_unqual_name()
        if last_unqual_name:
            for idx, img_path in enumerate(self.image_list):
                img_name = os.path.basename(img_path)
                if img_name == last_unqual_name:
                    self.current_index = idx + 1
                    if self.current_index >= len(self.image_list):
                        self.current_index = len(self.image_list) - 1
                    messagebox.showinfo("断点续检",
                                        f"检测到历史质检记录，最后不合格：{last_unqual_name}\n将从该文件的下一个位置开始质检！")
                    break

        # 重置状态，启用按钮
        self.checked_set.clear()
        self.prev_btn.config(state=tk.NORMAL)
        self.qual_btn.config(state=tk.NORMAL)
        self.unqual_btn.config(state=tk.NORMAL)
        self._show_current_image()
        self.root.lift()
        self.folder_label.config(text=f"文件夹：{os.path.basename(self.folder_path)}")
        self.status_label.config(text=f"状态：共{len(self.image_list)}张图片，待质检")

    def _init_excel(self):
        """初始化Excel，创建表头（不合格图片记录）"""
        if not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "不合格图片记录"
            ws["A1"] = "不合格图片名称（含后缀）"
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions["A"].width = 50
            # 表头标红，更醒目
            ws["A1"].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            wb.save(self.excel_path)
            print(f"Excel记录文件已创建：{self.excel_path}")

    def _show_current_image(self):
        """显示当前图片和进度"""
        if 0 <= self.current_index < len(self.image_list):
            image_path = self.image_list[self.current_index]
            file_name = os.path.basename(image_path)
            self.file_name_label.config(
                text=f"当前图片：{file_name} | 进度：{self.current_index + 1}/{len(self.image_list)}")
            self.status_label.config(text=f"状态：共{len(self.image_list)}张，已判定{len(self.checked_set)}张")
            self._resize_image()
        else:
            # 质检完成
            self.image_label.config(image="", text="✅ 所有图片质检完成！")
            self.file_name_label.config(text="当前图片：无")
            self.status_label.config(text=f"状态：质检完成，共判定{len(self.checked_set)}张，不合格已记录至Excel")
            self.prev_btn.config(state=tk.DISABLED)
            self.qual_btn.config(state=tk.DISABLED)
            self.unqual_btn.config(state=tk.DISABLED)
            messagebox.showinfo("完成", f"所有图片质检结束！\n不合格图片已记录至：{self.excel_path}")

    def _add_to_excel(self, file_name):
        """将不合格图片名称写入Excel，自动去重"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            # 读取已有记录，避免重复
            existing_names = []
            for i in range(2, ws.max_row + 1):
                val = ws.cell(row=i, column=1).value
                if val and val.strip():
                    existing_names.append(val.strip())
            # 仅写入新的不合格记录
            if file_name not in existing_names:
                ws.append([file_name])
                # 不合格记录行标浅红
                ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE",
                                                                     fill_type="solid")
                wb.save(self.excel_path)
            wb.close()
        except Exception as e:
            messagebox.showerror("错误", f"Excel写入失败：{str(e)}\n请关闭Excel后重试")
            print(f"Excel写入详情：{e}")

    def mark_qualified(self):
        """标记为合格：仅切换下一张，不记录"""
        if 0 <= self.current_index < len(self.image_list):
            file_name = os.path.basename(self.image_list[self.current_index])
            self.checked_set.add(file_name)
            self.status_label.config(text=f"✅ {file_name} 合格 | 已判定{len(self.checked_set)}张")
            self.current_index += 1
            self._show_current_image()

    def mark_unqualified(self):
        """标记为不合格：记录到Excel，切换下一张"""
        if 0 <= self.current_index < len(self.image_list):
            file_name = os.path.basename(self.image_list[self.current_index])
            self._add_to_excel(file_name)  # 写入不合格记录
            self.checked_set.add(file_name)
            self.status_label.config(text=f"❌ {file_name} 已记录 | 已判定{len(self.checked_set)}张")
            self.current_index += 1
            self._show_current_image()

    def prev_image(self):
        """回看上一张"""
        if self.current_index > 0:
            self.current_index -= 1
            self._show_current_image()


if __name__ == "__main__":
    # 自动安装依赖
    try:
        from PIL import Image
        import openpyxl
    except ImportError:
        print("正在自动安装依赖库，请稍候...")
        os.system("pip install pillow openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple")
        from PIL import Image
        import openpyxl

    # 解决中文乱码
    root = tk.Tk()
    try:
        root.option_add("*Font", "宋体 9")
    except:
        pass
    app = ImageCheckGUI(root)
    root.mainloop()